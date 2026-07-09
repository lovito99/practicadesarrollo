from io import BytesIO

from flask import jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from modelos.reservaModelo import actualizarEstadoReserva, buscarReservaPorId, listarReporteReservas
from seguridad.jwtServicio import jwtRequerido


ESTADOS_RESERVA = ("Pendiente", "Confirmada", "Cancelada", "Completada")


ENCABEZADOS = [
    "ID",
    "Cliente",
    "Correo",
    "Album",
    "Artista",
    "Reserva",
    "Evento",
    "Cant.",
    "Estado",
    "Total",
]


def formatearMonto(valor):
    return f"S/ {float(valor or 0):.2f}"


def filasReporte(reservas):
    return [
        [
            reserva["idReserva"],
            reserva["cliente"],
            reserva.get("correoCliente") or "-",
            reserva["tituloAlbum"],
            reserva["nombreArtista"],
            reserva.get("fechaReserva") or "-",
            reserva.get("fechaEvento") or "-",
            reserva["cantidad"],
            reserva["estado"],
            formatearMonto(reserva.get("montoTotal")),
        ]
        for reserva in reservas
    ]


def construirResumenReservas(reservas):
    return {
        "totalReservas": len(reservas),
        "totalEntradas": sum(int(reserva.get("cantidad") or 0) for reserva in reservas),
        "totalImporte": sum(float(reserva.get("montoTotal") or 0) for reserva in reservas),
    }


@jwtRequerido
def obtenerReporteReservas():
    try:
        reservas = listarReporteReservas(request.usuario["idUsuario"])

        return jsonify(
            {
                "ok": True,
                "reservas": reservas,
                "resumen": construirResumenReservas(reservas),
            }
        ), 200
    except Exception as error:
        return jsonify({"ok": False, "mensaje": str(error)}), 500


@jwtRequerido
def cambiarEstadoReserva(idReserva):
    datos = request.json or {}
    estado = str(datos.get("estado", "")).strip().capitalize()

    if estado not in ESTADOS_RESERVA:
        return jsonify(
            {
                "ok": False,
                "mensaje": "Estado no valido.",
                "estadosPermitidos": ESTADOS_RESERVA,
            }
        ), 400

    try:
        filasAfectadas = actualizarEstadoReserva(idReserva, estado, request.usuario["idUsuario"])
        if filasAfectadas == 0:
            return jsonify({"ok": False, "mensaje": "Reserva no encontrada."}), 404

        reserva = buscarReservaPorId(idReserva, request.usuario["idUsuario"])
        return jsonify(
            {
                "ok": True,
                "mensaje": "Estado de reserva actualizado correctamente.",
                "reserva": reserva,
            }
        ), 200
    except Exception as error:
        return jsonify({"ok": False, "mensaje": str(error)}), 500


@jwtRequerido
def exportarReporteReservasExcel():
    try:
        reservas = listarReporteReservas(request.usuario["idUsuario"])
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Reservas"

        hoja.append(["Reporte de reservas realizadas"])
        hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ENCABEZADOS))
        hoja["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        hoja["A1"].fill = PatternFill("solid", fgColor="16646F")
        hoja["A1"].alignment = Alignment(horizontal="center")

        hoja.append(ENCABEZADOS)
        for celda in hoja[2]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="7A5B31")
            celda.alignment = Alignment(horizontal="center")

        for fila in filasReporte(reservas):
            hoja.append(fila)

        anchos = [10, 22, 30, 26, 22, 14, 14, 10, 14, 14]
        for indice, ancho in enumerate(anchos, start=1):
            hoja.column_dimensions[hoja.cell(row=2, column=indice).column_letter].width = ancho

        for fila in hoja.iter_rows(min_row=3):
            for celda in fila:
                celda.alignment = Alignment(vertical="top")

        archivo = BytesIO()
        libro.save(archivo)
        archivo.seek(0)

        return send_file(
            archivo,
            as_attachment=True,
            download_name="reporte_reservas.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as error:
        return jsonify({"ok": False, "mensaje": str(error)}), 500


@jwtRequerido
def exportarReporteReservasPdf():
    try:
        reservas = listarReporteReservas(request.usuario["idUsuario"])
        archivo = BytesIO()
        documento = SimpleDocTemplate(
            archivo,
            pagesize=landscape(letter),
            rightMargin=24,
            leftMargin=24,
            topMargin=28,
            bottomMargin=24,
        )
        estilos = getSampleStyleSheet()
        elementos = [
            Paragraph("Reporte de reservas realizadas", estilos["Title"]),
            Spacer(1, 12),
        ]

        tabla = Table([ENCABEZADOS] + filasReporte(reservas), repeatRows=1)
        tabla.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16646F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5CDC2")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F3EF")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (0, -1), "CENTER"),
                    ("ALIGN", (7, 1), (9, -1), "RIGHT"),
                ]
            )
        )

        elementos.append(tabla)
        documento.build(elementos)
        archivo.seek(0)

        return send_file(
            archivo,
            as_attachment=True,
            download_name="reporte_reservas.pdf",
            mimetype="application/pdf",
        )
    except Exception as error:
        return jsonify({"ok": False, "mensaje": str(error)}), 500
